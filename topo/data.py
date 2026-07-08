"""Spreadsheet loading and boulder grouping.

Rows are one problem each; boulders are the unique `boulder` values, keyed in
insertion order. GPS/altitude/bearing/accuracy come from each boulder's photo
(EXIF), not the sheet.
"""
import csv
import json
import math
import os

from .exif import read_gps
from .lines import parse_line
from .style import BRAND, CAMERA_OFFSET_M, LINE_PALETTE, REFUGE


def _load_gallery(data_dir):
    """Read data/gallery.json (boulder name → [filename, ...]). Returns an
    empty dict if the file is missing or malformed."""
    path = os.path.join(data_dir, "gallery.json")
    try:
        with open(path) as f:
            obj = json.load(f)
        return obj if isinstance(obj, dict) else {}
    except Exception:
        return {}


def load_cluster_names(data_dir):
    """Read data/clusters.json (cluster letter → user-defined name). Returns
    an empty dict if the file is missing or malformed."""
    path = os.path.join(data_dir, "clusters.json")
    try:
        with open(path) as f:
            obj = json.load(f)
        return obj if isinstance(obj, dict) else {}
    except Exception:
        return {}


def load_cluster_assignments(data_dir):
    """Read data/cluster_assignments.json (boulder name → cluster letter).
    An entry here overrides the automatic longitude-gap clustering for that
    boulder. Returns an empty dict if the file is missing or malformed."""
    path = os.path.join(data_dir, "cluster_assignments.json")
    try:
        with open(path) as f:
            obj = json.load(f)
        return obj if isinstance(obj, dict) else {}
    except Exception:
        return {}


def _offset_by_bearing(lat, lon, bearing_deg, distance_m):
    """Shift (lat, lon) `distance_m` metres in the compass bearing direction
    (0° = north, clockwise). Uses the flat-earth approximation, which is
    fine for the ~4m nudges we care about here."""
    theta = math.radians(bearing_deg)
    dlat = distance_m * math.cos(theta) / 111320.0
    dlon = distance_m * math.sin(theta) / (111320.0 * math.cos(math.radians(lat)))
    return lat + dlat, lon + dlon


def load_rows(sheet_path):
    """Load rows from an .xlsx or .csv spreadsheet into a list of dicts."""
    rows = []
    if sheet_path.lower().endswith(".csv"):
        with open(sheet_path, newline="") as f:
            for row in csv.DictReader(f):
                rows.append(row)
    else:
        from openpyxl import load_workbook

        wb = load_workbook(sheet_path)
        ws = wb["Boulders"] if "Boulders" in wb.sheetnames else wb.active
        headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        for r in range(2, ws.max_row + 1):
            row = {headers[c - 1]: ws.cell(r, c).value for c in range(1, ws.max_column + 1)}
            if any(v not in (None, "") for v in row.values()):
                rows.append(row)
    return rows


def build_boulders(rows, photos_dir, gallery_dir=None):
    """Group rows into boulders keyed by boulder name; attach GPS + rendered photo."""
    # data/gallery.json lives one dir above photos_dir (both under data/).
    data_root = os.path.dirname(os.path.abspath(photos_dir))
    gallery_map = _load_gallery(data_root)
    gallery_dir = gallery_dir or os.path.join(data_root, "gallery")

    order = []
    groups = {}
    for row in rows:
        name = (row.get("boulder") or "").strip()
        if not name:
            continue
        if name not in groups:
            groups[name] = []
            order.append(name)
        groups[name].append(row)

    boulders = []
    for name in order:
        grp = groups[name]
        photo_file = str(grp[0].get("photo") or "").strip()
        photo_path = os.path.join(photos_dir, photo_file)
        problems = []
        for row in grp:
            grade = str(row.get("grade") or "").strip()
            problem_name = str(row.get("problem") or "").strip()
            no = row.get("no")
            try:
                no = int(no)
            except Exception:
                no = len(problems) + 1
            # Project is now a first-class boolean column. Fall back to the
            # legacy convention where "Project" lived in either the grade or
            # the problem-name cell so old CSVs still parse.
            project_cell = str(row.get("project") or "").strip().lower()
            if project_cell:
                project = project_cell in ("true", "1", "yes", "y")
            else:
                project = grade.lower() == "project" or problem_name.lower() == "project"
            color = "#000000" if project else LINE_PALETTE[(no - 1) % len(LINE_PALETTE)]
            problems.append(
                {
                    "no": no,
                    "name": problem_name,
                    "grade": grade or "–",
                    "notes": str(row.get("notes") or "").strip(),
                    "notes_fr": str(row.get("notes_fr") or "").strip(),
                    "line_pts": parse_line(row.get("line")),
                    "project": project,
                    "color": color,
                }
            )
        problems.sort(key=lambda p: p["no"])

        gps = None
        if os.path.exists(photo_path):
            try:
                gps = read_gps(photo_path)
            except Exception as e:
                print("  EXIF fail", photo_file, e)

        gallery_files = gallery_map.get(name, []) or []
        gallery_paths = [
            os.path.join(gallery_dir, f)
            for f in gallery_files
            if isinstance(f, str) and f
        ]
        b = {
            "name": name,
            "photo": photo_file,
            "photo_path": photo_path,
            "problems": problems,
            "gallery": [p for p in gallery_paths if os.path.exists(p)],
        }
        if gps:
            lat, lon = gps["lat"], gps["lon"]
            # If the photo has a compass bearing, nudge the recorded (camera)
            # position forward by CAMERA_OFFSET_M so it points at the boulder
            # itself instead of where the photographer was standing.
            if CAMERA_OFFSET_M and gps["bearing"] is not None:
                lat, lon = _offset_by_bearing(lat, lon, gps["bearing"], CAMERA_OFFSET_M)
            b.update(
                lat=lat,
                lon=lon,
                alt=gps["alt"],
                bearing=gps["bearing"],
                acc=gps["acc"],
            )
            b["alt_str"] = f"{gps['alt']:.0f} m" if gps["alt"] else "–"
            b["bearing_str"] = (
                f"{gps['bearing']:.0f}°" if gps["bearing"] is not None else "–"
            )
            b["_has_gps"] = True
        else:
            print(f"  WARNING: no GPS for boulder '{name}' (photo: {photo_file})")
            b.update(lat=REFUGE["lat"], lon=REFUGE["lon"], alt=None, bearing=None, acc=None)
            b["alt_str"] = "–"
            b["bearing_str"] = "–"
            b["_has_gps"] = False
        boulders.append(b)

    # Number boulders west → east (ascending longitude) so they follow the
    # river in that order. Boulders without real GPS fall back to CSV order
    # at the tail — we can't place them along the river without coordinates.
    with_gps = [b for b in boulders if b["_has_gps"]]
    without_gps = [b for b in boulders if not b["_has_gps"]]
    with_gps.sort(key=lambda b: b["lon"])
    ordered = with_gps + without_gps
    for i, b in enumerate(ordered, 1):
        b["id"] = i
    return ordered


def cluster_by_lon_gap(boulders, n_clusters=3, assignments=None):
    """Split boulders (that have real GPS) into `n_clusters` groups.

    If `assignments` (boulder-name → cluster-letter) is provided and any
    boulder is listed, those explicit assignments win — boulders are
    grouped by letter (west→east within each group). Any boulder without
    an assignment falls back to the automatic longitude-gap clustering
    used elsewhere.

    Returns (clusters, without_gps): clusters is a list of `n_clusters`
    lists, each ordered west→east; without_gps is a flat list.
    """
    with_gps = sorted(
        (b for b in boulders if b.get("_has_gps")),
        key=lambda b: b["lon"],
    )
    without_gps = [b for b in boulders if not b.get("_has_gps")]

    # Manual mode — group by the user's assignments, sorted west→east
    # within each cluster. Assumes assignments use A, B, C, ... letters.
    assignments = assignments or {}
    manual_used = any(assignments.get(b["name"]) for b in with_gps)
    if manual_used:
        letters = "ABCDEFGHIJKL"
        groups = {L: [] for L in letters[:n_clusters]}
        unassigned = []
        for b in with_gps:
            letter = (assignments.get(b["name"]) or "").strip().upper()
            if letter in groups:
                groups[letter].append(b)
            else:
                unassigned.append(b)
        # Fold any un-assigned GPS boulders into whichever cluster is
        # nearest (by mean longitude) so nothing goes missing.
        for b in unassigned:
            nearest = min(
                (L for L in groups if groups[L]),
                key=lambda L: abs(
                    b["lon"] - sum(x["lon"] for x in groups[L]) / len(groups[L])
                ),
                default="A",
            )
            groups[nearest].append(b)
        clusters = []
        for L in letters[:n_clusters]:
            clusters.append(sorted(groups[L], key=lambda b: b["lon"]))
        return clusters, without_gps

    if len(with_gps) <= n_clusters:
        return ([[b] for b in with_gps] +
                [[] for _ in range(n_clusters - len(with_gps))], without_gps)
    gaps = [
        (with_gps[k + 1]["lon"] - with_gps[k]["lon"], k)
        for k in range(len(with_gps) - 1)
    ]
    split_after = sorted(k for _, k in sorted(gaps, reverse=True)[: n_clusters - 1])
    clusters, start = [], 0
    for k in split_after:
        clusters.append(with_gps[start : k + 1])
        start = k + 1
    clusters.append(with_gps[start:])
    return clusters, without_gps
