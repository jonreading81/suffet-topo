#!/usr/bin/env python3
"""
Refuge du Suffet bouldering topo generator.

Turns a folder of GPS-tagged boulder photos + a spreadsheet of route data into:
  - a print-ready PDF topo
  - a standalone offline HTML map (IGN tiles + photos embedded; works with no signal)

Usage:
    python generate.py --input data --output output

Input folder layout:
    data/
      boulders.xlsx            (or .csv) with columns:
                               photo, boulder, no, problem, grade, notes, line
      photos/<the image files referenced in the `photo` column>

Notes:
  - GPS/altitude/bearing/accuracy are read from each photo's EXIF (not the sheet).
  - `line` is a list of "x,y" points as PERCENT of the image (from the annotator).
  - Needs internet at build time to fetch IGN map tiles.

This is an early, working starting point reconstructed from prototype code.
Expect to refactor (split into modules, add tests, a proper CLI).
"""
import argparse, base64, io, math, os, sys, csv, json
import urllib.request
from PIL import Image, ImageDraw, ImageFont

try:
    import pillow_heif
    pillow_heif.register_heif_opener()
except Exception:
    pass

# --------------------------------------------------------------------------
# CONFIG
# --------------------------------------------------------------------------
REFUGE = {"name": "Refuge du Suffet", "lat": 45.206031, "lon": 6.845828, "alt": 1690}

BRAND = {
    "blue":   "#004AAD",   # primary
    "teal":   "#6AB0AB",   # refuge marker / secondary
    "lav":    "#A096EF",   # projects
    "orange": "#E4572E",   # high-contrast line
    "amber":  "#E0A21B",
}
# problem line colours (graded); projects always use lavender + dashed
LINE_PALETTE = [BRAND["blue"], BRAND["orange"], BRAND["teal"], BRAND["lav"], BRAND["amber"]]

TILE_LAYERS = {
    "aerial": ("ORIMAGERY.ORTHOPHOTOS", "image/jpeg", "jpeg"),  # placeholder fixed below
}
IGN = "https://data.geopf.fr/wmts"
IGN_AERIAL = "ORTHOIMAGERY.ORTHOPHOTOS"
IGN_TOPO   = "GEOGRAPHICALGRIDSYSTEMS.PLANIGNV2"
ZOOM_MIN, ZOOM_MAX = 16, 18          # bundled zoom range for the offline map
ACCURACY_FLAG_M = 15                 # GPS accuracy above this is flagged low-confidence

HERE = os.path.dirname(os.path.abspath(__file__))
ASSETS = os.path.join(HERE, "assets")
FONT_DIR = "/usr/share/fonts/truetype/dejavu"

# --------------------------------------------------------------------------
# helpers: colour + fonts
# --------------------------------------------------------------------------
def hx(c):
    c = c.lstrip("#")
    return tuple(int(c[i:i+2], 16) for i in (0, 2, 4))

def font(size, bold=True):
    name = "DejaVuSans-Bold.ttf" if bold else "DejaVuSans.ttf"
    try:
        return ImageFont.truetype(os.path.join(FONT_DIR, name), size)
    except Exception:
        return ImageFont.load_default()

# --------------------------------------------------------------------------
# EXIF GPS
# --------------------------------------------------------------------------
def read_gps(path):
    """Return dict(lat, lon, alt, bearing, acc) or None."""
    from PIL.ExifTags import GPSTAGS
    img = Image.open(path)
    exif = img.getexif()
    gi = exif.get_ifd(0x8825)
    if not gi:
        return None
    g = {GPSTAGS.get(k, k): v for k, v in gi.items()}
    if "GPSLatitude" not in g or "GPSLongitude" not in g:
        return None
    def dd(v, ref):
        d = float(v[0]) + float(v[1]) / 60 + float(v[2]) / 3600
        return -d if ref in ("S", "W") else d
    return {
        "lat": dd(g["GPSLatitude"], g.get("GPSLatitudeRef", "N")),
        "lon": dd(g["GPSLongitude"], g.get("GPSLongitudeRef", "E")),
        "alt": float(g["GPSAltitude"]) if "GPSAltitude" in g else None,
        "bearing": float(g["GPSImgDirection"]) if "GPSImgDirection" in g else None,
        "acc": float(g["GPSHPositioningError"]) if "GPSHPositioningError" in g else None,
    }

# --------------------------------------------------------------------------
# geometry: line smoothing + drawing
# --------------------------------------------------------------------------
def parse_line(s):
    if not s:
        return []
    return [tuple(float(v) for v in pt.split(",")) for pt in str(s).split()]

def catmull(pts, steps=20):
    if len(pts) < 2:
        return pts
    out = []
    for i in range(len(pts) - 1):
        p0 = pts[i-1] if i > 0 else pts[i]
        p1, p2 = pts[i], pts[i+1]
        p3 = pts[i+2] if i+2 < len(pts) else p2
        for s in range(steps + 1):
            t = s / steps; t2 = t*t; t3 = t2*t
            x = 0.5*((2*p1[0]) + (-p0[0]+p2[0])*t + (2*p0[0]-5*p1[0]+4*p2[0]-p3[0])*t2 + (-p0[0]+3*p1[0]-3*p2[0]+p3[0])*t3)
            y = 0.5*((2*p1[1]) + (-p0[1]+p2[1])*t + (2*p0[1]-5*p1[1]+4*p2[1]-p3[1])*t2 + (-p0[1]+3*p1[1]-3*p2[1]+p3[1])*t3)
            out.append((x, y))
    return out

def draw_dashed(draw, pts, fill, width, dash, gap):
    on = True; carry = 0
    for i in range(len(pts) - 1):
        x1, y1 = pts[i]; x2, y2 = pts[i+1]
        seg = math.hypot(x2-x1, y2-y1)
        if seg == 0:
            continue
        d = 0
        while d < seg:
            step = (dash if on else gap) - carry
            end = min(d + step, seg)
            if on:
                draw.line([(x1+(x2-x1)*d/seg, y1+(y2-y1)*d/seg),
                           (x1+(x2-x1)*end/seg, y1+(y2-y1)*end/seg)], fill=fill, width=width)
            if end - d >= step:
                on = not on; carry = 0
            else:
                carry += end - d
            d = end

def render_boulder_photo(photo_path, problems, max_px=1100):
    """Draw all problem lines onto the photo in the shared style. Returns PIL.Image."""
    img = Image.open(photo_path).convert("RGB")
    W, H = img.size; base = min(W, H)
    d = ImageDraw.Draw(img, "RGBA")
    lw = max(2, int(base * 0.009)); halo = max(4, int(base * 0.016)); r = int(base * 0.024)
    fnt = font(int(r * 1.2))
    for p in problems:
        pts = [(x/100*W, y/100*H) for x, y in p["line_pts"]]
        if len(pts) < 2:
            continue
        cu = catmull(pts); col = hx(p["color"])
        if p["project"]:
            draw_dashed(d, cu, (255,255,255,210), halo, int(base*0.03), int(base*0.02))
            draw_dashed(d, cu, col+(255,), lw, int(base*0.03), int(base*0.02))
        else:
            d.line(cu, fill=(255,255,255,210), width=halo, joint="curve")
            d.line(cu, fill=col+(255,), width=lw, joint="curve")
        x0, y0 = pts[0]
        if p["project"]:
            d.ellipse([x0-r,y0-r,x0+r,y0+r], fill=(255,255,255,235), outline=col+(255,), width=max(3,int(base*0.008)))
            tcol = col+(255,)
        else:
            d.ellipse([x0-r,y0-r,x0+r,y0+r], fill=col+(255,), outline=(255,255,255,255), width=max(2,int(base*0.006)))
            tcol = (255,255,255,255)
        no = str(p["no"]); tb = d.textbbox((0,0), no, font=fnt)
        d.text((x0-(tb[2]-tb[0])/2, y0-(tb[3]-tb[1])/2-tb[1]), no, fill=tcol, font=fnt)
    if max_px:
        img.thumbnail((max_px, int(max_px*1.4)))
    return img

# --------------------------------------------------------------------------
# IGN tiles
# --------------------------------------------------------------------------
_TILE_CACHE = {}
def tile_url(layer, z, x, y, fmt):
    return (f"{IGN}?SERVICE=WMTS&REQUEST=GetTile&VERSION=1.0.0&LAYER={layer}"
            f"&STYLE=normal&TILEMATRIXSET=PM&TILEMATRIX={z}&TILEROW={y}&TILECOL={x}&FORMAT={fmt}")

def fetch_tile(layer, z, x, y, fmt):
    key = (layer, z, x, y)
    if key in _TILE_CACHE:
        return _TILE_CACHE[key]
    req = urllib.request.Request(tile_url(layer, z, x, y, fmt), headers={"User-Agent": "Mozilla/5.0"})
    data = urllib.request.urlopen(req, timeout=25).read()
    _TILE_CACHE[key] = data
    return data

def deg2tile(lat, lon, z):
    n = 2**z; la = math.radians(lat)
    return (int((lon+180)/360*n),
            int((1 - math.log(math.tan(la)+1/math.cos(la))/math.pi)/2*n))

def global_px(lat, lon, z):
    n = 2**z; la = math.radians(lat)
    return ((lon+180)/360*n*256,
            (1 - math.log(math.tan(la)+1/math.cos(la))/math.pi)/2*n*256)

def stitch_map(points, refuge, out_path, z=17, Wc=1200, Hc=720):
    """points: list of dicts with lat/lon/label. Draws refuge + boulder pins."""
    all_pts = [(refuge["lat"], refuge["lon"])] + [(p["lat"], p["lon"]) for p in points]
    cx = sum(global_px(la, lo, z)[0] for la, lo in all_pts) / len(all_pts)
    cy = sum(global_px(la, lo, z)[1] for la, lo in all_pts) / len(all_pts)
    ox, oy = cx - Wc/2, cy - Hc/2
    canvas = Image.new("RGB", (Wc, Hc), (200, 200, 200))
    c0, c1 = int(ox//256), int((ox+Wc)//256); r0, r1 = int(oy//256), int((oy+Hc)//256)
    for col in range(c0, c1+1):
        for row in range(r0, r1+1):
            try:
                t = Image.open(io.BytesIO(fetch_tile(IGN_AERIAL, z, col, row, "image/jpeg"))).convert("RGB")
                canvas.paste(t, (int(col*256-ox), int(row*256-oy)))
            except Exception as e:
                print("  tile miss", z, col, row, e)
    d = ImageDraw.Draw(canvas, "RGBA")
    f1, f2 = font(22), font(17)
    teal, blue = hx(BRAND["teal"]), hx(BRAND["blue"])
    # refuge marker (teal square, house)
    sx, sy = global_px(refuge["lat"], refuge["lon"], z); sx -= ox; sy -= oy
    w = 17
    d.rectangle([sx-w,sy-w,sx+w,sy+w], fill=teal+(255,), outline=(255,255,255,255), width=3)
    d.polygon([(sx-9,sy-1),(sx,sy-10),(sx+9,sy-1)], fill=(255,255,255,255))
    d.rectangle([sx-6,sy-1,sx+6,sy+8], fill=(255,255,255,255))
    lb = d.textbbox((0,0), refuge["name"], font=f2)
    d.rectangle([sx+w+3,sy-11,sx+w+(lb[2]-lb[0])+11,sy+13], fill=(255,255,255,230))
    d.text((sx+w+7,sy-9), refuge["name"], fill=teal+(255,), font=f2)
    # boulder pins
    for i, p in enumerate(points):
        bx, by = global_px(p["lat"], p["lon"], z); bx -= ox; by -= oy
        rr = 15; label = p.get("label", str(i+1))
        d.ellipse([bx-rr,by-rr,bx+rr,by+rr], fill=blue+(255,), outline=(255,255,255,255), width=3)
        tb = d.textbbox((0,0), label, font=f2)
        d.text((bx-(tb[2]-tb[0])/2, by-(tb[3]-tb[1])/2-tb[1]), label, fill=(255,255,255,255), font=f2)
        nm = p.get("name", "")
        if nm:
            lb = d.textbbox((0,0), nm, font=f2)
            d.rectangle([bx+rr+2,by-11,bx+rr+(lb[2]-lb[0])+10,by+13], fill=(255,255,255,220))
            d.text((bx+rr+6,by-9), nm, fill=blue+(255,), font=f2)
    # scale bar + N + attribution
    mpp = 156543.03392*math.cos(math.radians(all_pts[0][0]))/(2**z)
    seg = 100/mpp; bxs, bys = 30, Hc-40
    d.rectangle([bxs-6,bys-24,bxs+seg+14,bys+16], fill=(0,0,0,130))
    d.line([(bxs,bys),(bxs+seg,bys)], fill=(255,255,255,255), width=3)
    for xx in (bxs, bxs+seg):
        d.line([(xx,bys-5),(xx,bys+5)], fill=(255,255,255,255), width=3)
    d.text((bxs,bys-22), "100 m", fill=(255,255,255,255), font=f2)
    d.text((Wc-40,18), "N", fill=(255,255,255,255), font=f1)
    d.line([(Wc-33,58),(Wc-33,28)], fill=(255,255,255,255), width=3)
    d.polygon([(Wc-33,24),(Wc-38,34),(Wc-28,34)], fill=(255,255,255,255))
    at = "\u00a9 IGN / G\u00e9oplateforme"; ab = d.textbbox((0,0), at, font=f2)
    d.rectangle([Wc-(ab[2]-ab[0])-14,Hc-26,Wc,Hc], fill=(0,0,0,130))
    d.text((Wc-(ab[2]-ab[0])-8,Hc-24), at, fill=(255,255,255,255), font=f2)
    canvas.save(out_path, "JPEG", quality=88)

def bundle_tiles(bbox):
    """bbox = (latmin, lonmin, latmax, lonmax). Returns {'layer/z/x/y': dataURI}."""
    latmin, lonmin, latmax, lonmax = bbox
    tiles = {}
    for key, layer, fmt, ext in [("aerial", IGN_AERIAL, "image/jpeg", "jpeg"),
                                 ("topo", IGN_TOPO, "image/png", "png")]:
        for z in range(ZOOM_MIN, ZOOM_MAX+1):
            x0, y1 = deg2tile(latmin, lonmin, z); x1, y0 = deg2tile(latmax, lonmax, z)
            for x in range(min(x0,x1), max(x0,x1)+1):
                for y in range(min(y0,y1), max(y0,y1)+1):
                    try:
                        data = fetch_tile(layer, z, x, y, fmt)
                        tiles[f"{key}/{z}/{x}/{y}"] = f"data:{fmt};base64," + base64.b64encode(data).decode()
                    except Exception as e:
                        print("  bundle miss", key, z, x, y, e)
    return tiles

# --------------------------------------------------------------------------
# PDF
# --------------------------------------------------------------------------
def build_pdf(boulders, out_path):
    from reportlab.lib.pagesizes import A4
    from reportlab.pdfgen import canvas as pdfcanvas
    from reportlab.lib.utils import ImageReader
    from reportlab.lib.colors import HexColor
    W, H = A4
    c = pdfcanvas.Canvas(out_path, pagesize=A4)
    BLUE = HexColor(BRAND["blue"]); TEAL = HexColor(BRAND["teal"]); LAV = HexColor(BRAND["lav"])
    ORANGE = HexColor(BRAND["orange"]); INK = HexColor("#1d1d1b"); MUT = HexColor("#6b6b66")
    LINE = HexColor("#dbe1ec"); CARD = HexColor("#eef2f8"); AMBER = HexColor("#B5751A"); WARN = HexColor("#FAEEDA")
    M = 42; SERIF = "Times-Bold"
    logo = ImageReader(os.path.join(ASSETS, "logo_white.png"))

    def header(t, s):
        c.setFillColor(BLUE); c.rect(0, H-96, W, 96, fill=1, stroke=0)
        lw2 = 42*806/306; c.drawImage(logo, W-M-lw2, H-96+(96-42)/2, lw2, 42, mask="auto")
        c.setFillColor(HexColor("#ffffff")); c.setFont(SERIF, 21); c.drawString(M, H-52, t)
        c.setFont("Helvetica", 10.5); c.setFillColor(HexColor("#cfe0ea")); c.drawString(M, H-72, s)

    def footer(p):
        c.setStrokeColor(LINE); c.setLineWidth(0.5); c.line(M, 40, W-M, 40)
        c.setFont("Helvetica", 8.5); c.setFillColor(MUT)
        c.drawString(M, 30, "Refuge du Suffet boulders \u00b7 built from photo GPS + spreadsheet")
        c.drawRightString(W-M, 30, str(p))

    # page 1: getting there map
    header("Refuge du Suffet boulders", "Haute-Maurienne \u00b7 getting there")
    cw = W-2*M; mtop = H-116
    im = Image.open(os.path.join(os.path.dirname(out_path), "_map.jpg")); iw, ih = im.size; hm = cw*ih/iw
    c.drawImage(ImageReader(os.path.join(os.path.dirname(out_path), "_map.jpg")), M, mtop-hm, cw, hm)
    c.setStrokeColor(LINE); c.setLineWidth(0.5); c.rect(M, mtop-hm, cw, hm, fill=0, stroke=1)
    c.setFont("Helvetica-Oblique", 7.5); c.setFillColor(MUT)
    c.drawString(M, mtop-hm-12, "IGN aerial \u00b7 the refuge and the boulders")
    footer(1); c.showPage()

    # one detail page per boulder
    pg = 2
    for b in boulders:
        header(b["name"], f"{b['lat']:.5f}\u00b0N, {b['lon']:.5f}\u00b0E  \u00b7  {b.get('alt_str','')}")
        im = Image.open(b["_render"]); iw, ih = im.size
        pw = 250; ph = pw*ih/iw; px = M; py = H-130-ph
        c.drawImage(ImageReader(b["_render"]), px, py, pw, ph)
        c.setStrokeColor(LINE); c.setLineWidth(0.5); c.rect(px, py, pw, ph, fill=0, stroke=1)
        rx = M+pw+26; rw = W-M-rx; yr = H-130; mh = 140
        c.setFillColor(CARD); c.roundRect(rx, yr-mh, rw, mh, 8, fill=1, stroke=0)
        c.setFillColor(BLUE); c.setFont(SERIF, 12); c.drawString(rx+14, yr-20, "Location")
        def meta(l, v, yy):
            c.setFont("Helvetica", 8.8); c.setFillColor(MUT); c.drawString(rx+14, yy, l)
            c.setFont("Helvetica", 9.5); c.setFillColor(INK); c.drawRightString(rx+rw-14, yy, v)
        meta("Latitude", f"{b['lat']:.5f}\u00b0N", yr-42)
        meta("Longitude", f"{b['lon']:.5f}\u00b0E", yr-60)
        meta("Altitude", b.get("alt_str", "\u2013"), yr-78)
        meta("Photo bearing", b.get("bearing_str", "\u2013"), yr-96)
        if b.get("acc") is not None:
            flagged = b["acc"] >= ACCURACY_FLAG_M
            c.setFillColor(WARN if flagged else CARD)
            c.roundRect(rx+14, yr-mh+14, rw-28, 18, 9, fill=1, stroke=0)
            c.setFillColor(AMBER if flagged else MUT); c.setFont("Helvetica-Bold", 8.5)
            msg = ("\u26a0  GPS \u00b1%d m \u2014 low confidence, verify on map" % b["acc"]) if flagged else ("GPS \u00b1%d m" % b["acc"])
            c.drawString(rx+22, yr-mh+20, msg)
        yp = yr-mh-26; c.setFillColor(INK); c.setFont(SERIF, 13); c.drawString(rx, yp, "Problems")
        yp -= 6; c.setStrokeColor(BLUE); c.setLineWidth(1.2); c.line(rx, yp, rx+rw, yp); yp -= 20
        for p in b["problems"]:
            r = 8; col = HexColor(p["color"])
            if p["project"]:
                c.setFillColor(HexColor("#ffffff")); c.setStrokeColor(col); c.setLineWidth(2)
                c.circle(rx+r, yp+3, r, fill=1, stroke=1); c.setFillColor(col)
            else:
                c.setFillColor(col); c.circle(rx+r, yp+3, r, fill=1, stroke=0); c.setFillColor(HexColor("#ffffff"))
            c.setFont("Helvetica-Bold", 9); c.drawCentredString(rx+r, yp, str(p["no"]))
            c.setFillColor(INK); c.setFont("Helvetica-Bold", 10.5); c.drawString(rx+24, yp, p["name"])
            c.setFillColor(LAV if p["project"] else BLUE); c.setFont("Helvetica-Bold", 10)
            c.drawRightString(rx+rw, yp, p["grade"])
            c.setFillColor(MUT); c.setFont("Helvetica", 9); ln = ""; ly = yp-15
            for wd in (p["notes"] or "").split():
                if c.stringWidth(ln+" "+wd, "Helvetica", 9) < rw-24:
                    ln = (ln+" "+wd).strip()
                else:
                    c.drawString(rx+24, ly, ln); ly -= 13; ln = wd
            c.drawString(rx+24, ly, ln)
            yp = ly-18
            c.setStrokeColor(LINE); c.setLineWidth(0.5); c.line(rx, yp+8, rx+rw, yp+8)
        footer(pg); c.showPage(); pg += 1
    c.setTitle("Refuge du Suffet boulders"); c.save()

# --------------------------------------------------------------------------
# offline HTML
# --------------------------------------------------------------------------
HTML_TEMPLATE = None  # loaded from build_html to keep braces out of f-strings

def build_html(boulders, refuge, tiles, bbox, out_path):
    css = open(os.path.join(ASSETS, "vendor", "leaflet.css")).read()
    js = open(os.path.join(ASSETS, "vendor", "leaflet.js")).read()
    logo = open(os.path.join(ASSETS, "logo_white.svg")).read()
    data = {"suffet": refuge, "boulders": []}
    for b in boulders:
        data["boulders"].append({
            "id": b["id"], "name": b["name"], "lat": b["lat"], "lon": b["lon"],
            "alt": b.get("alt"), "acc": b.get("acc"), "bearing": b.get("bearing_str", "\u2013"),
            "photo": b["_photo_uri"],
            "problems": [{"n": p["no"], "name": p["name"], "grade": p["grade"],
                          "color": p["color"], "project": p["project"], "beta": p["notes"]} for p in b["problems"]],
        })
    DATA = json.dumps(data)
    latmin, lonmin, latmax, lonmax = bbox
    # content bounds = tight around points
    pts = [(refuge["lat"], refuge["lon"])] + [(b["lat"], b["lon"]) for b in boulders]
    clatmin = min(p[0] for p in pts) - 0.0005; clatmax = max(p[0] for p in pts) + 0.0005
    clonmin = min(p[1] for p in pts) - 0.0008; clonmax = max(p[1] for p in pts) + 0.0008

    html = _HTML.replace("__CSS__", css).replace("__JS__", js) \
                .replace("__TILES__", json.dumps(tiles)).replace("__DATA__", DATA) \
                .replace("__LOGO__", logo) \
                .replace("__TB__", f"[[{latmin},{lonmin}],[{latmax},{lonmax}]]") \
                .replace("__CB__", f"[[{clatmin},{clonmin}],[{clatmax},{clonmax}]]") \
                .replace("__ZMIN__", str(ZOOM_MIN)).replace("__ZMAX__", str(ZOOM_MAX))
    open(out_path, "w").write(html)

# offline HTML template (brand-styled). Placeholders: __CSS__ __JS__ __TILES__ __DATA__ __LOGO__ __TB__ __CB__ __ZMIN__ __ZMAX__
_HTML = r"""<!DOCTYPE html>
<html lang="en"><head><meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1, maximum-scale=1, user-scalable=no">
<title>Refuge du Suffet boulders</title>
<style>__CSS__</style>
<style>
*{box-sizing:border-box} html,body{margin:0;height:100%;font-family:-apple-system,BlinkMacSystemFont,"Segoe UI",Roboto,sans-serif;color:#1d1d1b}
#app{position:fixed;inset:0;display:flex;flex-direction:column}
#topbar{background:#004AAD;color:#fff;padding:10px 14px;z-index:1200;flex:0 0 auto;display:flex;align-items:center;justify-content:space-between;gap:12px}
#topbar .tt h1{margin:0;font-size:17px;font-weight:700;font-family:Georgia,"Times New Roman",serif}
#topbar .tt p{margin:2px 0 0;font-size:11.5px;color:#cfe0ea}
#topbar .logo svg{height:34px;width:auto;display:block}
#map{flex:1 1 auto;background:#e8edf4;z-index:1}
.pin{display:flex;align-items:center;justify-content:center;width:30px;height:30px;border-radius:50%;background:#004AAD;color:#fff;font-weight:700;font-size:13px;border:2.5px solid #fff;box-shadow:0 1px 4px rgba(0,0,0,.4)}
.pin.sel{background:#E4572E;transform:scale(1.12)}
.ref{display:flex;align-items:center;justify-content:center;width:30px;height:30px;border-radius:6px;background:#6AB0AB;color:#fff;border:2.5px solid #fff;box-shadow:0 1px 4px rgba(0,0,0,.4);font-size:16px}
#sheet{position:absolute;left:0;right:0;bottom:0;background:#fff;z-index:1100;border-top-left-radius:16px;border-top-right-radius:16px;box-shadow:0 -4px 20px rgba(0,0,0,.18);transform:translateY(102%);transition:transform .28s ease;max-height:64%;overflow-y:auto;padding:0 16px 20px}
#sheet.open{transform:translateY(0)}
.grab{width:38px;height:4px;background:#c8d2e0;border-radius:2px;margin:8px auto 4px}
.sheethd{display:flex;align-items:baseline;justify-content:space-between;gap:8px;padding:4px 0 6px;position:sticky;top:0;background:#fff}
.sheethd h2{margin:0;font-size:18px;font-weight:700;font-family:Georgia,serif;color:#004AAD}
.close{border:none;background:#eef2f8;border-radius:8px;font-size:18px;line-height:1;padding:6px 10px;cursor:pointer}
.metarow{display:flex;flex-wrap:wrap;gap:6px;margin:2px 0 10px}
.chip{font-size:11px;background:#e7eef7;color:#33506e;border-radius:6px;padding:3px 8px}
.chip.warn{background:#FAEEDA;color:#8a5a12;font-weight:600}
.photo{width:100%;border-radius:10px;border:.5px solid #dbe1ec;margin-bottom:10px}
.nophoto{width:100%;aspect-ratio:4/3;border-radius:10px;border:1px dashed #c3ccdb;display:flex;align-items:center;justify-content:center;color:#98a1b3;font-size:13px;margin-bottom:10px;background:#f7f9fc}
.prob{border-top:.5px solid #eaeef4;padding:9px 0}.prob:first-of-type{border-top:none}
.probhd{display:flex;align-items:center;gap:8px}
.dot{width:20px;height:20px;border-radius:50%;color:#fff;font-size:11px;font-weight:700;display:flex;align-items:center;justify-content:center;flex:0 0 auto}
.dot.proj{background:#fff!important;border:2px solid #A096EF;color:#A096EF}
.pname{font-weight:600;font-size:14px}
.pgrade{margin-left:auto;font-weight:700;color:#004AAD;font-size:14px}
.pgrade.proj{color:#A096EF}
.pbeta{font-size:12.5px;color:#555;margin:3px 0 0 28px;line-height:1.45}
.legend{font-size:11px;color:#8890a0;font-style:italic;margin-top:8px}
#hint{position:absolute;top:10px;left:50%;transform:translateX(-50%);z-index:1000;background:rgba(255,255,255,.92);border-radius:20px;padding:6px 14px;font-size:12px;color:#33506e;box-shadow:0 1px 6px rgba(0,0,0,.15)}
.leaflet-control-attribution{font-size:9px}
</style></head><body><div id="app">
<div id="topbar"><div class="tt"><h1>Refuge du Suffet boulders</h1><p>Haute-Maurienne \u00b7 tap a marker \u00b7 works offline</p></div><div class="logo">__LOGO__</div></div>
<div id="map"><div id="hint">Tap a boulder for problems + beta</div></div></div>
<div id="sheet"><div class="grab"></div><div id="sheetbody"></div></div>
<script>__JS__</script>
<script>
var TILES=__TILES__, DATA=__DATA__, EMPTY="data:image/gif;base64,R0lGODlhAQABAAAAACwAAAAAAQABAAA=";
var Off=L.TileLayer.extend({getTileUrl:function(c){return TILES[this.options.k+"/"+c.z+"/"+c.x+"/"+c.y]||EMPTY;}});
var tileBounds=L.latLngBounds(__TB__), contentBounds=L.latLngBounds(__CB__);
var map=L.map('map',{minZoom:__ZMIN__,maxZoom:__ZMAX__,maxBounds:tileBounds,maxBoundsViscosity:1.0,zoomControl:true});
var aerial=new Off('',{k:'aerial',minZoom:__ZMIN__,maxZoom:__ZMAX__,bounds:tileBounds,attribution:'\u00a9 IGN / G\u00e9oplateforme'});
var topo=new Off('',{k:'topo',minZoom:__ZMIN__,maxZoom:__ZMAX__,bounds:tileBounds,attribution:'\u00a9 IGN / G\u00e9oplateforme'});
aerial.addTo(map); L.control.layers({'Aerial':aerial,'Topo':topo},null,{collapsed:false}).addTo(map);
map.fitBounds(contentBounds);
var s=DATA.suffet;
L.marker([s.lat,s.lon],{icon:L.divIcon({className:'',html:'<div class="ref">\u2302</div>',iconSize:[30,30],iconAnchor:[15,15]})}).addTo(map).bindTooltip(s.name+' \u00b7 '+s.alt+' m',{direction:'right',offset:[14,0]});
DATA.boulders.forEach(function(b){
 var m=L.marker([b.lat,b.lon],{icon:L.divIcon({className:'',html:'<div class="pin" id="pin'+b.id+'">'+b.id+'</div>',iconSize:[30,30],iconAnchor:[15,15]})}).addTo(map);
 m.bindTooltip(b.name,{direction:'right',offset:[14,0]}); m.on('click',function(){openB(b);});});
var sheet=document.getElementById('sheet'),body=document.getElementById('sheetbody');
function openB(b){
 document.querySelectorAll('.pin').forEach(function(p){p.classList.remove('sel')});
 var pe=document.getElementById('pin'+b.id); if(pe)pe.classList.add('sel');
 var acc=(b.acc!=null&&b.acc>=15)?'<span class="chip warn">\u26a0 GPS \u00b1'+b.acc+' m</span>':(b.acc!=null?'<span class="chip">GPS \u00b1'+b.acc+' m</span>':'');
 var h='<div class="sheethd"><h2>'+b.name+'</h2><button class="close" onclick="closeSheet()">\u00d7</button></div>';
 h+='<div class="metarow"><span class="chip">'+b.lat.toFixed(5)+'\u00b0N, '+b.lon.toFixed(5)+'\u00b0E</span>'+(b.alt?'<span class="chip">'+b.alt+' m</span>':'')+'<span class="chip">bearing '+b.bearing+'</span>'+acc+'</div>';
 h+=b.photo?('<img class="photo" src="'+b.photo+'">'):'<div class="nophoto">photo added from your upload</div>';
 var hasP=false;
 b.problems.forEach(function(p){var pr=p.project;if(pr)hasP=true;
   h+='<div class="prob"><div class="probhd"><div class="dot'+(pr?' proj':'')+'" style="'+(pr?'':'background:'+p.color)+'">'+p.n+'</div><span class="pname">'+p.name+'</span><span class="pgrade'+(pr?' proj':'')+'">'+p.grade+'</span></div><div class="pbeta">'+p.beta+'</div></div>';});
 if(hasP) h+='<div class="legend">Dashed line + open marker = project (unclimbed).</div>';
 body.innerHTML=h; sheet.classList.add('open'); map.panTo([b.lat,b.lon]);
}
function closeSheet(){sheet.classList.remove('open');document.querySelectorAll('.pin').forEach(function(p){p.classList.remove('sel')});}
map.on('click',closeSheet);
</script></body></html>"""

# --------------------------------------------------------------------------
# data loading
# --------------------------------------------------------------------------
def load_rows(sheet_path):
    rows = []
    if sheet_path.lower().endswith(".csv"):
        with open(sheet_path, newline="") as f:
            for row in csv.DictReader(f):
                rows.append(row)
    else:
        from openpyxl import load_workbook
        wb = load_workbook(sheet_path)
        ws = wb["Boulders"] if "Boulders" in wb.sheetnames else wb.active
        headers = [ws.cell(1, c).value for c in range(1, ws.max_column+1)]
        for r in range(2, ws.max_row+1):
            row = {headers[c-1]: ws.cell(r, c).value for c in range(1, ws.max_column+1)}
            if any(v not in (None, "") for v in row.values()):
                rows.append(row)
    return rows

def build_boulders(rows, photos_dir):
    """Group rows into boulders keyed by boulder name; attach GPS + rendered photo."""
    order = []; groups = {}
    for row in rows:
        name = (row.get("boulder") or "").strip()
        if not name:
            continue
        if name not in groups:
            groups[name] = []; order.append(name)
        groups[name].append(row)
    boulders = []
    for i, name in enumerate(order, 1):
        grp = groups[name]
        photo_file = str(grp[0].get("photo") or "").strip()
        photo_path = os.path.join(photos_dir, photo_file)
        problems = []
        for row in grp:
            grade = str(row.get("grade") or "").strip()
            no = row.get("no")
            try: no = int(no)
            except Exception: no = len(problems)+1
            project = grade.lower() == "project"
            color = BRAND["lav"] if project else LINE_PALETTE[(no-1) % len(LINE_PALETTE)]
            problems.append({
                "no": no, "name": str(row.get("problem") or "").strip(),
                "grade": grade or "\u2013", "notes": str(row.get("notes") or "").strip(),
                "line_pts": parse_line(row.get("line")), "project": project, "color": color,
            })
        problems.sort(key=lambda p: p["no"])
        gps = None
        if os.path.exists(photo_path):
            try: gps = read_gps(photo_path)
            except Exception as e: print("  EXIF fail", photo_file, e)
        b = {"id": i, "name": name, "photo": photo_file, "photo_path": photo_path,
             "problems": problems}
        if gps:
            b.update(lat=gps["lat"], lon=gps["lon"], alt=gps["alt"], bearing=gps["bearing"], acc=gps["acc"])
            b["alt_str"] = f"{gps['alt']:.0f} m" if gps["alt"] else "\u2013"
            b["bearing_str"] = f"{gps['bearing']:.0f}\u00b0" if gps["bearing"] is not None else "\u2013"
        else:
            print(f"  WARNING: no GPS for boulder '{name}' (photo: {photo_file})")
            b.update(lat=REFUGE["lat"], lon=REFUGE["lon"], alt=None, bearing=None, acc=None)
            b["alt_str"] = "\u2013"; b["bearing_str"] = "\u2013"
        boulders.append(b)
    return boulders

# --------------------------------------------------------------------------
# main
# --------------------------------------------------------------------------
def main():
    ap = argparse.ArgumentParser(description="Generate the Refuge du Suffet bouldering topo (PDF + offline HTML).")
    ap.add_argument("--input", default="data", help="input folder (spreadsheet + photos/)")
    ap.add_argument("--output", default="output", help="output folder")
    args = ap.parse_args()

    photos_dir = os.path.join(args.input, "photos")
    sheets = [f for f in os.listdir(args.input)
              if f.lower().endswith((".xlsx", ".csv")) and not f.startswith("~")]
    if not sheets:
        sys.exit("No .xlsx/.csv spreadsheet found in " + args.input)
    sheet_path = os.path.join(args.input, sheets[0])
    print("Spreadsheet:", sheet_path)

    os.makedirs(args.output, exist_ok=True)
    rows = load_rows(sheet_path)
    boulders = build_boulders(rows, photos_dir)
    print(f"{len(boulders)} boulder(s), {sum(len(b['problems']) for b in boulders)} problem(s)")

    # render each boulder photo with its lines (full-res for PDF, small for HTML)
    for b in boulders:
        if os.path.exists(b["photo_path"]):
            big = render_boulder_photo(b["photo_path"], b["problems"], max_px=1100)
            b["_render"] = os.path.join(args.output, f"_boulder_{b['id']}.jpg")
            big.save(b["_render"], "JPEG", quality=88)
            small = render_boulder_photo(b["photo_path"], b["problems"], max_px=560)
            buf = io.BytesIO(); small.save(buf, "JPEG", quality=72)
            b["_photo_uri"] = "data:image/jpeg;base64," + base64.b64encode(buf.getvalue()).decode()
        else:
            b["_render"] = None; b["_photo_uri"] = None
            print("  photo missing:", b["photo_path"])

    # getting-there map
    map_points = [{"lat": b["lat"], "lon": b["lon"], "name": b["name"], "label": str(b["id"])} for b in boulders]
    stitch_map(map_points, REFUGE, os.path.join(args.output, "_map.jpg"))

    # bounding box for offline tiles (around refuge + boulders, padded)
    lats = [REFUGE["lat"]] + [b["lat"] for b in boulders]
    lons = [REFUGE["lon"]] + [b["lon"] for b in boulders]
    padlat, padlon = 0.004, 0.006
    bbox = (min(lats)-padlat, min(lons)-padlon, max(lats)+padlat, max(lons)+padlon)

    print("Building PDF...")
    build_pdf(boulders, os.path.join(args.output, "refuge-du-suffet-boulders.pdf"))
    print("Bundling offline tiles (needs internet)...")
    tiles = bundle_tiles(bbox)
    print(f"  {len(tiles)} tiles bundled")
    print("Building offline HTML...")
    build_html(boulders, REFUGE, tiles, bbox, os.path.join(args.output, "refuge-du-suffet-boulders.html"))
    print("Done ->", args.output)

if __name__ == "__main__":
    main()
